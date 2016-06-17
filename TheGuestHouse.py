import openpyxl
from Person import *


wb = openpyxl.load_workbook("theguesthouse.xlsx")

room_301 = wb.get_sheet_by_name('301')
room_302 = wb.get_sheet_by_name('302')
room_303 = wb.get_sheet_by_name('303')
room_304 = wb.get_sheet_by_name('304')
room_305 = wb.get_sheet_by_name('305')
room_306 = wb.get_sheet_by_name('306')
room_307 = wb.get_sheet_by_name('307')
room_308 = wb.get_sheet_by_name('308')
room_309 = wb.get_sheet_by_name('309')
room_401 = wb.get_sheet_by_name('401')
room_402 = wb.get_sheet_by_name('402')
room_403 = wb.get_sheet_by_name('403')
room_404 = wb.get_sheet_by_name('404')
room_405 = wb.get_sheet_by_name('405')
room_406 = wb.get_sheet_by_name('406')
room_407 = wb.get_sheet_by_name('407')
room_408 = wb.get_sheet_by_name('408')
room_409 = wb.get_sheet_by_name('409')
room_410 = wb.get_sheet_by_name('410')
room_310 = wb.get_sheet_by_name('310')

sheet_list_thirdfloor = [room_301,room_302,room_303,room_304,room_305,room_306,room_307,room_308,room_309,room_310]
sheet_list_fourthfloor = [room_401,room_402,room_403,room_404,room_405,room_406,room_407,room_408,room_409,room_410]

def addPerson(room_number, name, gender, phone_number, ssn, start_date):
    person = Person(name,gender,phone_number,ssn,start_date)
    if room_number < 400:
        index = (room_number % 100) - 1
        rowNum = 6
        while sheet_list_thirdfloor[index].cell(row=rowNum, column=1).value != None:
            rowNum += 1
        sheet_list_thirdfloor[index].cell(row=rowNum, column=1).value = person.name
        sheet_list_thirdfloor[index].cell(row=rowNum, column=2).value = person.ssn
        sheet_list_thirdfloor[index].cell(row=rowNum, column=3).value = person.phone_number
        sheet_list_thirdfloor[index].cell(row=rowNum, column=4).value = person.gender
        sheet_list_thirdfloor[index].cell(row=rowNum, column=5).value = person.start_date


    elif room_number > 400:
        index = (room_number % 100) - 1
        rowNum = 6
        while sheet_list_fourthfloor[index].cell(row=rowNum, column=1).value != None:
            rowNum += 1
        sheet_list_fourthfloor[index].cell(row=rowNum, column=1).value = person.name
        sheet_list_fourthfloor[index].cell(row=rowNum, column=2).value = person.ssn
        sheet_list_fourthfloor[index].cell(row=rowNum, column=3).value = person.phone_number
        sheet_list_fourthfloor[index].cell(row=rowNum, column=4).value = person.gender
        sheet_list_fourthfloor[index].cell(row=rowNum, column=5).value = person.start_date

    wb.save("theguesthouse.xlsx")

def main():


    room_num = int(input("호실: "))
    name = str(input("이름: "))
    gender = str(input("성별: "))
    phone = str(input("Phone: "))
    ssn = str(input("SSN: "))
    start_date = str(input("Start date: "))

    addPerson(room_num,name,gender,phone,ssn,start_date)

main()


